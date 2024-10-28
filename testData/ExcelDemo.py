import openpyxl
book=openpyxl.load_workbook("C:/Users/manoj.b/PycharmProjects/PythonSelfFramework/testData/Book1.xlsx")
sheet=book.active
cell=sheet.cell(row=1,column=2)
#To fetch data
print(cell.value)
#To insert data
sheet.cell(row=2,column=2).value="manoj"
print(sheet.cell(row=2,column=2).value)
#to get max row count
print(sheet.max_row)
#To get max column count
print(sheet.max_column)
#other way to get data from sheet
print(sheet['A2'].value)
print("======================================================")

#To Get All The Data From The execl in loop
for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
       print(sheet.cell(row=i,column=j).value)


print("======================================================")
#If Want to get data of Password row oly but need all the column
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "Password":
         for j in range(1,sheet.max_column+1):
              print(sheet.cell(row=i,column=j).value)